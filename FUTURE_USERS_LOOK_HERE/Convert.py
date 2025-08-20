import json
import os
from openpyxl import load_workbook

script_dir = os.path.dirname(os.path.abspath(__file__))
xlsx_file = os.path.join(script_dir, "tutors.xlsx")
json_file = os.path.join(script_dir, "../backend/data/tutors.json")

wb = load_workbook(xlsx_file)
sheet = wb.active

tutors_list = []
seen_ids = set()  # Track IDs we've already processed

for row in sheet.iter_rows(min_row=2):
    # Extract and clean Student ID
    id_value = row[2].value
    if id_value is None:
        student_id = "Unknown_ID"
    else:
        try:
            student_id = str(int(float(id_value)))
        except (ValueError, TypeError):
            student_id = str(id_value).strip()
    
    # Skip if ID has already been used
    if student_id in seen_ids:
        continue
    seen_ids.add(student_id)

    # Extract Full Name
    full_name = row[3].value.strip() if row[3].value else "Unknown Name"
    
    # Extract and clean Grade
    grade_value = row[4].value
    if grade_value is None:
        grade = "Unknown Grade"
    else:
        try:
            grade = str(int(float(grade_value)))
        except (ValueError, TypeError):
            grade = str(grade_value).strip()
    
    # Subject columns: Math, Science, Social Studies, Foreign Languages, Misc, English, What Books can you tutor?
    subject_cells = row[5:12]

    subjects = []
    for cell in subject_cells:
        if cell.value:
            val = str(cell.value).strip()
            # Split by double spaces or commas if applicable
            if "  " in val:
                parts = val.split("  ")
            elif "," in val:
                parts = [p.strip() for p in val.split(",")]
            else:
                parts = [val]
            subjects.extend(parts)

    # Remove any empty strings
    subjects = [s for s in subjects if s]

    tutor_json = {
        "id": student_id,
        "name": full_name,
        "available": False,
        "photo": f"{full_name}.jpeg",
        "grade": f"{grade}",
        "subjects": subjects
    }

    tutors_list.append(tutor_json)

with open(json_file, "w") as f:
    json.dump(tutors_list, f, indent=2)

print(f"Saved {len(tutors_list)} tutors to {json_file}")